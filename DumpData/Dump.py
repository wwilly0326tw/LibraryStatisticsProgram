from openpyxl import load_workbook
import mysql.connector
import DBconfig as DBconfig

conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                               host=DBconfig.host)
cur = conn.cursor()

outputFile = open("r.txt", 'w+', encoding='UTF-8')
outputFile.write('Access_Type')
outputFile.write('\t')
outputFile.write('nSubscribe')
outputFile.write('\t')
outputFile.write('nFree')
outputFile.write('\t')
outputFile.write('Themes')
outputFile.write('\t')
outputFile.write('Departments')
outputFile.write('\t')
outputFile.write('Colleges')
outputFile.write('\t')
outputFile.write('Targets')
outputFile.write('\n')

def dump():
    wb = load_workbook(filename="scopus.xlsx")
    ws = wb[wb.sheetnames[0]]
    dataStr = 'A2:A' + str(ws.max_row)
    for row in ws[dataStr]:
        scopusID = row[0].value
        cur.execute("Select sfxID from support where scopusID = " + str(scopusID))
        sfxIDList = cur.fetchall()

        if sfxIDList is not None and len(sfxIDList) is not 0:
            sfxIDstr = ""
            themeIDStr = ""
            targetNameStr = ""
            nPaid = 0
            nFree = 0
            for sfxID in sfxIDList:
                # 判別是否有付費
                try:
                    cur.execute("SELECT isfree from sfx where id = " + str(sfxID[0]))
                    if cur.fetchone()[0] == 0:
                        nPaid += 1
                    else:
                        nFree += 1
                except Exception as err:
                    continue
                # 串sfxID
                sfxIDstr += str(sfxID[0])
                sfxIDstr += ","
                # 蒐集sfx所對應的主題id
                try:
                    cur.execute("SELECT tid from relation_sfx_theme where sfxid = " + str(sfxID[0]))
                    resultOfTid = cur.fetchall()
                    for Tid in resultOfTid:
                        themeIDStr += str(Tid[0])
                        themeIDStr += ","
                except Exception as err:
                    continue
                # 蒐集sfx所對應的target name
                try:
                    cur.execute("SELECT TargetPublicName from sfx where id=" + str(sfxID[0]))
                    targetName = cur.fetchone()[0]
                    targetNameStr += "\""
                    targetNameStr += targetName
                    targetNameStr += "\""
                    targetNameStr += ","
                except Exception as err:
                    continue
            if nPaid - nFree == nPaid:
                outputFile.write("Subscribed")
            elif nFree - nPaid == nFree:
                outputFile.write("Free")
            else:
                outputFile.write("Mixed")
            outputFile.write("\t")
            outputFile.write(str(nPaid))
            outputFile.write("\t")
            outputFile.write(str(nFree))
            outputFile.write("\t")
            # 用主題串成的ID去找對應的科系及院別
            outputResult(sfxIDstr, themeIDStr, targetNameStr)
        else:
            outputFile.write("Not_Found")
            outputFile.write("\n")
    return

def outputResult(sfxIDList, themeIDList, targetNameList):
    # 尋找主題
    themeStr = ""
    try:
        cur.execute("SELECT name from theme where tid in (" + themeIDList[0: len(themeIDList) - 1] + ")")
        resultOfTheme = cur.fetchall()
        for theme in resultOfTheme:
            themeStr += theme[0]
            themeStr += "|"
        themeStr = themeStr[0: len(themeStr) - 1]
        outputFile.write(themeStr)
        outputFile.write('\t')
    except Exception as err:
        pass
    # 尋找sfx所對應的科系
    departmentStr = ""
    collegeID = ""
    try:
        cur.execute(
            "SELECT name, cid from department where did in (select did from relation_sfx_department where sfxid in (" + sfxIDList[
                                                                                                                        0: len(
                                                                                                                            sfxIDList) - 1] + "))")
        resultOfDepart = cur.fetchall()
        for depart in resultOfDepart:
            collegeID += str(depart[1])
            collegeID += ","
            departmentStr += depart[0]
            departmentStr += "|"
        departmentStr = departmentStr[0: len(departmentStr) - 1]
        outputFile.write(departmentStr)
        outputFile.write('\t')
    except Exception as err:
        pass
    # 尋找科系所對應的院
    collegeStr = ""
    try:
        cur.execute("SELECT name from college where cid in (" + collegeID[0: len(collegeID) - 1] + ")")
        resultOfCollege = cur.fetchall()
        for college in resultOfCollege:
            collegeStr += college[0]
            collegeStr += "|"
        collegeStr = collegeStr[0: len(collegeStr) - 1]
        outputFile.write(collegeStr)
        outputFile.write('\t')
    except Exception as err:
        pass
    # 尋找對應的Target
    targetNameStr = ""
    try:
        cur.execute("SELECT name from target where name in (" + targetNameList[0: len(targetNameList) - 1] + ")")
        resultOfTarget = cur.fetchall()
        for target in resultOfTarget:
            targetNameStr += target[0]
            targetNameStr += "|"
        targetNameStr = targetNameStr[0: len(targetNameStr) - 1]
        outputFile.write(targetNameStr)
        outputFile.write('\n')
    except Exception as err:
        pass

if __name__ == '__main__':
    dump()