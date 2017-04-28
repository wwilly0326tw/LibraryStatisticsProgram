import mysql.connector
import Program.DBconfig as DBconfig
import sys
from openpyxl import load_workbook

try:
    conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                   host=DBconfig.host)
    cur = conn.cursor()
except Exception as err:
    sys.exit(-1)

id = 92925
wb = load_workbook(filename="../../Data/SFX/20170421SFX_adv_both_both.xlsx")
ws = wb[wb.sheetnames[0]]
sqlStmt = "update sfx set ISBN = %s eISBN = %s where id = %d"
for row in ws['A1:AO' + str(ws.max_row)]:
    ISBN = row[23].value
    eISBN = row[24].value
    try:
        cur.execute(sqlStmt % (ISBN, eISBN, id))
        conn.commit()
    except Exception as err:
        print (err)
        continue
