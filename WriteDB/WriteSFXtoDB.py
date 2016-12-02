from openpyxl import load_workbook
import mysql.connector
import logging
import logging.config


def writeSFX2DB(filename=""):
    if filename is "":
        print ("Please Input a file.")
        return False

    logging.config.fileConfig("../logger.conf")
    logger = logging.getLogger("root")

    try:
        conn = mysql.connector.connect(user='user', password='zxc105', database='librarystatisticsdata', host='localhost')
        cur = conn.cursor()
    except Exception as err:
        logger.error(err)
        return False

    sqlStmt = "insert into sfx(SortableTitle,Title,TitleNon‐FilingCharacter,ISSN,ObjectID,TargetPublicName,Threshold,Eissn,AbbreviatedTitle,TargetServiceType,LCCN,ObjectPortfolioID,856‐u,856‐y,856‐a,245_h,TargetID,TargetServiceID,ObjectPortfolio_ID,Categories,LocalAttribute,ISBN,eISBN,Publisher,PlaceofPublication,DateofPublication,ObjectType,ActivationstatusfortheDEFAULTinstitute,InstituteID,InstituteName,InstituteAvailability,Language,MainTitle,FullOriginalTitle,AdditionalISBNs,AdditionaleISBNs,Author,Owner, isFree) values ("

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in ws['A1:AL' + str(ws.max_row)]:
        valStr = ""
        for col in row:
            if col.value is None:
                valStr += 'NULL'
            else:
                valStr += "\""
                valStr += str(col.value).replace("\"", "'").replace("\\", "")
                valStr += "\""
            valStr += ", "
        dbStr = row[5].value
        if dbStr.find('free') is not -1 or dbStr.find('Free') is not -1 or dbStr.find('Open Access') is not -1:
            valStr += "1"
        else:
            valStr += "0"
        valStr += ")"
        try:
            cur.execute(sqlStmt + valStr)
            conn.commit()
        except Exception as err:
            print(sqlStmt + valStr)
            logger.error(err)
            conn.rollback()
            continue


if __name__ == '__main__':
    writeSFX2DB('..\..\Data\SFX\sfx.xlsx')


