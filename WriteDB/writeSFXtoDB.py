from openpyxl import load_workbook
import DBconfig as DBconfig
import mysql.connector
import logging
import logging.config

debug = 0

logging.config.fileConfig("../logger.conf")
logger = logging.getLogger("root")

try:
    conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                   host=DBconfig.host)
    cur = conn.cursor()
except Exception as err:
    logger.error(err)

""" 此程式用於將sfx資料寫入資料庫中 """
def writeSFX2DB(filename="", year=""):
    if filename is "":
        print ("Please Input a file.")
        return False
    sqlStmt = "insert into sfx(SortableTitle,Title,TitleNon‐FilingCharacter,ISSN,ObjectID,TargetPublicName,Threshold,Eissn,AbbreviatedTitle,TargetServiceType,LCCN,ObjectPortfolioID,856‐u,856‐y,856‐a,245_h,LocalThreshold,GlobalThreshold,TargetID,TargetServiceID,ObjectPortfolio_ID,Categories,LocalAttribute,ISBN,eISBN,Publisher,PlaceofPublication,DateofPublication,ObjectType,ActivationstatusfortheDEFAULTinstitute,InstituteID,InstituteName,InstituteAvailability,Language,MainTitle,FullOriginalTitle,AdditionalISBNs,AdditionaleISBNs,Author,Owner,THRESHOLD_LOCAL, isFree, year) values ("

    wb = load_workbook(filename=filename)
    ws = wb[wb.sheetnames[0]]

    for row in ws['A2:AO' + str(ws.max_row)]:
        valStr = ""
        ISSN = row[3].value
        if ISSN is not None:
            ISSN = ISSN.replace("-", "")
            row[3].value = ISSN
        eISSN = row[7].value
        if eISSN is not None:
            eISSN = eISSN.replace("-", "")
            row[7].value = eISSN
        checkTargetinDB(row[5].value)
        for col in row:
            if col.value is None:
                valStr += 'NULL'
            else:
                valStr += "\""
                valStr += str(col.value).replace("\"", "'").replace("\\", "")
                valStr += "\""
            valStr += ", "
        dbStr = row[5].value
        if dbStr is not None:
            if dbStr.find('free') is not -1 or dbStr.find('Free') is not -1 or dbStr.find('Open Access') is not -1:
                valStr += "1"
            else:
                valStr += "0"
        valStr += ", "
        valStr += year
        valStr += ")"
        try:
            if debug:
                print(sqlStmt + valStr)
            cur.execute(sqlStmt + valStr)
            conn.commit()
        except Exception as err:
            print (err)
            logger.error(err)
            conn.rollback()
            continue

def checkTargetinDB(target):
    if target is not None:
        stmt = "Select count(*) from target  where name = \"" + target + "\""
        try:
            if debug:
                print (stmt)
            cur.execute(stmt)
            exist = cur.fetchone()
            if not exist[0]:
                stmt = "Insert into target(name) values(\"" + target + "\")"
                if debug:
                    print (stmt)
                cur.execute(stmt)
        except Exception as err:
            print (err)
            logger.error(err)
            conn.rollback()

if __name__ == '__main__':
    writeSFX2DB('../../Data/SFX/testsfx.xlsx', "2018")
    # writeSFX2DB('./testdata.xlsx', "2018")

