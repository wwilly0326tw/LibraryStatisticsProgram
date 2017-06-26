from openpyxl import load_workbook
import DBconfig as DBconfig
import mysql.connector
import logging
import logging.config

""" 程式用途為將主題名稱寫入資料庫，以及建立主題與對應科系的關聯 """

def relate_Theme_Department(filename="", year=""):
    if filename is "":
        print ("Please Input a file.")
        return False

    logging.config.fileConfig("../logger.conf")
    logger = logging.getLogger("root")

    try:
        conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                       host=DBconfig.host)
        cur = conn.cursor()
        selectStmt = "select did, name from department"
        cur.execute(selectStmt)
        resultOfDep = cur.fetchall()
    except Exception as err:
        logger.error(err)
        return False

    insertStmt = "insert into theme(name, year) values (%s, %s)"

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in ws['A2:I' + str(ws.max_row)]:
        valStr = ""
        valStr += row[5].value[row[5].value.find(' ') + 1:]
        valStr += "/"
        valStr += row[3].value[row[3].value.find(' ') + 1:]
        valStr += " - "
        valStr += row[6].value[row[6].value.find(' ') + 1:]
        valStr += "/"
        valStr += row[4].value[row[4].value.find(' ') + 1:]
        try:
            cur.execute(insertStmt, (valStr, year))
            conn.commit()
        except Exception as err:
            print(insertStmt + valStr)
            logger.error(err)
            conn.rollback()
            continue
        departmentID = []
        departmentCol = str(row[8].value)
        for department in resultOfDep:
            if departmentCol.find(department[1]) != -1:
                departmentID.append(department[0])
        if departmentID:
            insertRelStmt = "insert into relation_theme_department(tid, did) values(" + str(cur.lastrowid) + ", "
            for ID in departmentID:
                try:
                    cur.execute(insertRelStmt + str(ID) + ")")
                    conn.commit()
                except Exception as err:
                    print(insertRelStmt + str(ID) + ")")
                    logger.error(err)
                    conn.rollback()
                    continue
            print (departmentID)


if __name__ == '__main__':
    relate_Theme_Department('..\..\Data\SFX\\1051219sfx主題分類-學院系所對照.xlsx')


