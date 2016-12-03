# -*-coding: utf-8-*-
import sys
import time
import Program.DBconfig as DBconfig
import mysql.connector
import logging
import logging.config
from termcolor.termcolor import colored
from openpyxl import Workbook
from openpyxl import load_workbook


PISSN = []
EISSN = []

"""從sfx中取出PISSN EISSN"""


def getISSN():
    print(colored('Reading SFX-ISSN...', 'blue', attrs=['bold', 'underline']))
    wb = load_workbook(filename='./Data/sfxISSN.xlsx', read_only=True)
    # wb = load_workbook(filename = 'test.xlsx', read_only=True)
    ws = wb[wb.sheetnames[0]]

    PISSN_str = 'A1:A' + str(ws.max_row)
    EISSN_str = 'B1:B' + str(ws.max_row)

    for row in ws[PISSN_str]:
        # print (cell.value)
        PISSN.append(row[0].value)
    for row in ws[EISSN_str]:
        # print (cell.value)
        EISSN.append(row[0].value)
    print(colored('Reading SFX-ISSN... ', 'blue', attrs=['bold', 'underline']) + colored('Completed.', 'yellow'))


"""比對PISSN EISSN 是否有相同的，若相同回傳其index"""


def cmpISSN(ISSN):

    try:
        conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                       host=DBconfig.host)
        cur = conn.cursor()
    except Exception as err:
        logger.error(err)
        return False

    try:
        cur.execute("SELECT id from sfx where ISSN = '" + ISSN + "'")
        return cur.fetchall()
    except Exception as err:
        logger.error(err)
        return None


def traceback(err):
    now = time.strftime('%H:%M:%S', time.localtime(time.time()))
    tb = sys.exc_info()[2]
    print(now, err, 'exception in line', tb.tb_lineno)


if __name__ == '__main__':
    logging.config.fileConfig("../logger.conf")
    logger = logging.getLogger("root")
    print(cmpISSN('0256-1891'))
