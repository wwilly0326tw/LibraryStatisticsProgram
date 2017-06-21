# -*- coding: utf-8 -*-
from CmpISSN.ISSN_Comp import cmpISSNISBN
from CmpInterval.cmpInterval import cmpInterval
from ISBNTransfer.ISBNTransfer import ISBN10to13
from time import gmtime, strftime
from openpyxl import load_workbook
from termcolor.termcolor import colored
import re
import sys
import logging
import logging.config
import mysql.connector
import DBconfig as DBconfig


"""此程式用於統合比對ISSN以及年卷期，回傳是否有購買此篇參考文獻"""
logging.config.fileConfig("./logger.conf")
logger = logging.getLogger("root")
FilePath = "./"
InputPath = "./"
toCommit = 0
debug = 1

try:
    conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                   host=DBconfig.host)
    cur = conn.cursor()
except Exception as err:
    logger.error(err)
    sys.exit(-1)

try:
    cur.execute("SELECT batchID FROM librarystatisticsdata.support order by batchID desc limit 1")
    batchID = cur.fetchone()[0]
    batchID += 1
except Exception as err:
    logger.error(err)
    sys.exit(-1)

try:
    outputFile = open(FilePath + str(batchID) + str(".txt"), 'w+', encoding='UTF-8')
    outputFile.write('Access_Type')
    outputFile.write('\t')
    outputFile.write('nSubscribe')
    outputFile.write('\t')
    outputFile.write('nFree')
    outputFile.write('\t')
    outputFile.write('Themes')
    outputFile.write('\t')
    outputFile.write('Departments')
    outputFile.write('\t')
    outputFile.write('Colleges')
    outputFile.write('\t')
    outputFile.write('Targets')
    outputFile.write('\n')
except Exception as err:
    logger.error(err)
    sys.exit(-1)


def main(filename="testdata.xlsx", year=""):
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    if filename is not "":
        threadCounter(0)
        wb = load_workbook(filename=InputPath + filename)
        ws = wb[wb.sheetnames[0]]
        dataStr = 'A2:J' + str(ws.max_row)
        for row in ws[dataStr]:
            if year == "":
                year = strftime("%Y", gmtime())
            isSupport = ""
            nPaid = 0
            nFree = 0
            sfxIDstr = ""
            themeIDStr = ""
            targetNameStr = ""
            if row[8].value is not None:
                row[8].value = re.sub("[^0-9Xx;]", "", str(row[8].value))
            ISSN = str(row[8].value)
            row[9].value = ISBN10to13(row[9].value) # 去除多餘字元
            ISBN = row[9].value

            if ISSN is not None:
                if ISBN is not None:
                    # ISSN 與 ISBN 皆有值時，若卷期都無值用ISBN
                    if row[4].value is None and row[5].value is None:
                        sfxIDList = cmpISSNISBN(ISBN = ISBN, year = year)  # 比對到ISBN的清單
                    else:
                        sfxIDList = cmpISSNISBN(ISSN = ISSN, year = year)  # 比對到ISSN的清單
                        ISBN = ""
                else:
                    # ISSN 有值 ISBN 無值
                    sfxIDList = cmpISSNISBN(ISSN = ISSN, year = year)  # 比對到ISSN的清單
            elif ISBN is not None:
                # ISSN　無值 ISBN 有值
                sfxIDList = cmpISSNISBN(ISBN = ISBN, year = year)  # 比對到ISBN的清單
            else:
                # ISSN ISBN 都無值
                outputFile.write("Null")
                outputFile.write('\n')
                continue
            scopusID = insertDB(row)  # 將scopus的資料insert到DB
            if scopusID == -1:
                outputFile.write("Data Error")
                outputFile.write('\n')
                continue
            if sfxIDList is not None and len(sfxIDList) is not 0:
                try:
                    cur.execute("SELECT year, volume, issue from scopus where id = " + str(scopusID))
                    YVI = cur.fetchone()
                except Exception as err:
                    logger.info('Select YVI from scopus error.')
                    logger.error(err)
                    continue
                for sfxID in sfxIDList:
                    try:
                        cur.execute("SELECT Threshold from sfx where id=" + str(sfxID[0]))
                        threshold = cur.fetchone()[0]
                    except Exception as err:
                        logger.info('Select threshold from sfx error.')
                        logger.error(err)
                        continue
                    # 若ISBN有值 代表此次比較是以ISBN為基準，不需要比對區間
                    if ISBN is not "" or cmpInterval(threshold, str(YVI[0]) + "." + str(YVI[1]) + "." + str(YVI[2])):
                        sfxIDstr += str(sfxID[0])
                        sfxIDstr += ","
                        isSupport = 1
                        try:
                            cur.execute(
                                "INSERT INTO support(sfxID, scopusID, batchID) values(" + str(sfxID[0]) + "," + str(
                                    scopusID) + "," + str(batchID) + ")")
                            if debug:
                                print("INSERT INTO support(sfxID, scopusID, batchID) values(" + str(sfxID[0]) + "," + str(
                                scopusID) + "," + str(batchID) + ")")
                            if toCommit:
                                conn.commit()
                        except Exception as err:
                            logger.info('Create relation in support error.')
                            logger.error(err)
                            continue
                        # 蒐集sfx所對應的主題id
                        try:
                            cur.execute("SELECT tid from relation_sfx_theme where sfxid = " + str(sfxID[0]))
                            resultOfTid = cur.fetchall()
                            for Tid in resultOfTid:
                                themeIDStr += str(Tid[0])
                                themeIDStr += ","
                        except Exception as err:
                            logger.info('Search themeID error.')
                            logger.error(err)
                            continue
                        # 蒐集sfx所對應的target name
                        try:
                            cur.execute("SELECT TargetPublicName from sfx where id=" + str(sfxID[0]))
                            targetName = cur.fetchone()[0]
                            targetNameStr += "\""
                            targetNameStr += targetName
                            targetNameStr += "\""
                            targetNameStr += ","
                        except Exception as err:
                            logger.info('Select TargetName from sfx error.')
                            logger.error(err)
                            continue
                        # 判別是否有付費
                        try:
                            cur.execute("SELECT isfree from sfx where id = " + str(sfxID[0]))
                            if cur.fetchone()[0] == 0:
                                nPaid += 1
                            else:
                                nFree += 1
                        except Exception as err:
                            logger.info('paid error.')
                            logger.error(err)
                            continue

                    else:
                        if debug:
                            print(threshold)
                            print(YVI)
                            print(colored('Interval not match.', 'red'))
                # 建立scopus與Target的關聯
                if targetNameStr is not "":
                    try:
                        cur.execute(
                            "SELECT tid from target where name in (" + targetNameStr[0: len(targetNameStr) - 1] + ")")
                        resultOfTarget = cur.fetchall()
                        for target in resultOfTarget:
                            try:
                                insertStr = "INSERT INTO relation_target_scopus(tid, scopusID, batchID) values(" + str(target[0]) + "," + str(
                                        scopusID) + "," + str(batchID) + ")"
                                cur.execute(insertStr)
                                if debug:
                                    print (insertStr)
                                if toCommit:
                                    conn.commit()
                            except Exception as err:
                                conn.rollback()
                                logger.info('INSERT relation_target_scopus error.')
                                logger.error(err)
                    except Exception as err:
                        if debug:
                            print (targetNameStr)
                        logger.info('Relate target and scopus error.')
                        logger.error(err)
            else:  # 未找到sfx 直接結束
                if debug:
                    print(colored('ISSN/ISBN not match.', 'red'))
                outputFile.write("Not_Found")
                outputFile.write('\n')
                continue

            if isSupport:
                insertTargetScore(targetNameStr, nPaid + nFree)
                if nPaid - nFree == nPaid:
                    outputFile.write("Subscribed")
                elif nFree - nPaid == nFree:
                    outputFile.write("Free")
                else:
                    outputFile.write("Mixed")
                outputFile.write("\t")
                outputFile.write(str(nPaid))
                outputFile.write("\t")
                outputFile.write(str(nFree))
                outputFile.write("\t")

                # 用主題串成的ID去找對應的科系及院別
                if themeIDStr is not "":
                    if debug:
                        print ("Themes ID string " + themeIDStr)
                        print(colored('Match.', 'yellow'))
                    outputResult(sfxIDstr, themeIDStr, targetNameStr)
            # 有比對到ISSN/ISBN但區間未比對到
            elif isSupport == "":
                outputFile.write("Not_Found")
            outputFile.write('\n')
    else:
        print('Please Input the File.')
        return
    print (batchID)
    outputFile.close()
    threadCounter(1)


def insertDB(row):
    sqlStmt = "INSERT INTO scopus(author_keyword, book_name, year, source_name, volume, issue, DOI, link, ISSN, ISBN) values("
    valStr = ""
    for col in row:
        if col.value is None:
            valStr += 'NULL'
        else:
            valStr += "\""
            valStr += str(col.value).replace("\"", "'").replace("\\", "")
            valStr += "\""
        valStr += ", "
    valStr = valStr[0: len(valStr) - 2]
    valStr += ")"
    try:
        cur.execute(sqlStmt + valStr)
        if debug:
            print(sqlStmt + valStr)
        if toCommit:
            conn.commit()
        # 修改卷期非數字的問題
        scoupusID = cur.lastrowid
    except Exception as err:
        if debug:
            print(sqlStmt + valStr)
        logger.error(err)
        conn.rollback()
        return -1
    modifyYVI(scoupusID)
    return scoupusID


def modifyYVI(scoupusID):
    try:
        cur.execute("SELECT year, volume, issue from scopus where id = " + str(scoupusID))
        result = cur.fetchone()
    except Exception as err:
        logger.info("Select YVI to modify error.")
        logger.error(err)
        return False
    updateFlag = 0
    tmpList = [result[0], result[1], result[2]]
    if result[0] is None or not re.match('^[0-9]+$', result[0]):
        tmpList[0] = 0
        updateFlag = 1
    if result[1] is None or not re.match('^[0-9]+$', result[1]):
        tmpList[1] = 0
        updateFlag = 1
    if result[2] is None or not re.match('^[0-9]+$', result[2]):
        tmpList[2] = 0
        updateFlag = 1
    if updateFlag:
        try:
            cur.execute("UPDATE scopus set year=" + str(tmpList[0]) + ",volume=" + str(tmpList[1]) + ", issue=" + str(
                tmpList[2]) + " where id = " + str(scoupusID))
            if debug:
                print("UPDATE scopus set year=" + str(tmpList[0]) + ",volume=" + str(tmpList[1]) + ", issue=" + str(
                tmpList[2]) + " where id = " + str(scoupusID))
            if toCommit:
                conn.commit()
        except Exception as err:
            logger.info('Update YVI error.')
            logger.error(err)
            return False


def outputResult(sfxIDList, themeIDList, targetNameList):
    # 尋找主題
    themeStr = ""
    try:
        cur.execute("SELECT name from theme where tid in (" + themeIDList[0: len(themeIDList) - 1] + ")")
        resultOfTheme = cur.fetchall()
        for theme in resultOfTheme:
            themeStr += theme[0]
            themeStr += "|"
        themeStr = themeStr[0: len(themeStr) - 1]
        if debug:
            print(themeStr)
        outputFile.write(themeStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search theme error.')
        logger.error(err)
    # 尋找sfx所對應的科系
    departmentStr = ""
    collegeID = ""
    try:
        cur.execute(
            "SELECT name, cid from department where did in (select did from relation_sfx_department where sfxid in (" + sfxIDList[
                                                                                                                        0: len(
                                                                                                                            sfxIDList) - 1] + "))")
        resultOfDepart = cur.fetchall()
        for depart in resultOfDepart:
            collegeID += str(depart[1])
            collegeID += ","
            departmentStr += depart[0]
            departmentStr += "|"
        departmentStr = departmentStr[0: len(departmentStr) - 1]
        if debug:
            print(departmentStr)
        outputFile.write(departmentStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search department error.')
        logger.error(err)
    # 尋找科系所對應的院
    collegeStr = ""
    try:
        cur.execute("SELECT name from college where cid in (" + collegeID[0: len(collegeID) - 1] + ")")
        resultOfCollege = cur.fetchall()
        for college in resultOfCollege:
            collegeStr += college[0]
            collegeStr += "|"
        collegeStr = collegeStr[0: len(collegeStr) - 1]
        if debug:
            print(collegeStr)
        outputFile.write(collegeStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search college error.')
        logger.error(err)
    # 尋找對應的Target
    targetNameStr = ""
    try:
        cur.execute("SELECT name from target where name in (" + targetNameList[0: len(targetNameList) - 1] + ")")
        resultOfTarget = cur.fetchall()
        for target in resultOfTarget:
            targetNameStr += target[0]
            targetNameStr += "|"
        targetNameStr = targetNameStr[0: len(targetNameStr) - 1]
        if debug:
            print(targetNameStr)
        outputFile.write(targetNameStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search target error.')
        logger.error(err)


def insertTargetScore(targetNameList, nSup):
    targetNameList = targetNameList[0:-1].split(",")
    nSup = 1 / float(nSup)
    for target in targetNameList:
        try:
            sql = "Update target set score = score + " + str(nSup) + " where name = " + target
            if debug:
                print (sql)
            cur.execute("Update target set score = score + " + str(nSup) + "where name = " + target)
            conn.commit()
        except Exception as err:
            conn.rollback()
            logger.info('Update score error.')
            logger.error(err)

def threadCounter(out):
    try:
        nThread = open('flag', 'r')
        flag = int(nThread.read())
        if out:
            flag -= 1
        else:
            flag += 1
        nThread = open("flag", 'w')
        nThread.write(str(flag))
        nThread.close()
    except Exception as err:
        logger.error(err)
        sys.exit(-1)

if __name__ == '__main__':
    # main(filename="../Data/scopus/scopus.xlsx")
    main()
