# -*- coding: utf-8 -*-
from CmpISSN.ISSN_Comp import cmpISSN
from CmpInterval.cmpInterval import cmpInterval
from openpyxl import load_workbook
from termcolor.termcolor import colored
import logging
import logging.config
import sys
import re
import DBconfig as DBconfig
import mysql.connector
import os

"""此程式用於統合比對ISSN以及年卷期，回傳是否有購買此篇參考文獻"""
# logging.config.fileConfig("./logger.conf")
logger = logging.getLogger("root")
debug = 0

try:
    conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                   host=DBconfig.host)
    cur = conn.cursor()
except Exception as err:
    logger.error(err)
    sys.exit(-1)

try:
    cur.execute("SELECT batchID FROM librarystatisticsdata.support order by batchID desc limit 1")
    batchID = cur.fetchone()[0]
    batchID += 1
except Exception as err:
    logger.error(err)
    sys.exit(-1)

try:
    outputFile = open("../result/" + str(batchID) + str(".txt"), 'w+', encoding = 'UTF-8')
except Exception as err:
    logger.error(err)
    sys.exit(-1)

def main(filename="testdata.xlsx"):
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    if filename is not "":
        wb = load_workbook(filename="../data/" + filename)
        ws = wb[wb.sheetnames[0]]
        dataStr = 'A2:J' + str(ws.max_row)
        for row in ws[dataStr]:
            isSupport = "nSupported"
            isPaid = ""
            themeIDStr = ""
            if row[8].value == "":
                continue
            sfxIDList = cmpISSN(row[8].value) # 比對到ISSN的清單
            scopusID = insertDB(row) # 將scopus的資料insert到DB
            if scopusID == -1:
                continue
            if sfxIDList is not None and len(sfxIDList) is not 0:
                try:
                    cur.execute("SELECT year, volume, issue from scopus where id = " + str(scopusID))
                    YVI = cur.fetchone()
                except Exception as err:
                    logger.info('Select YVI from scopus error.')
                    logger.error(err)
                    continue
                for sfxID in sfxIDList:
                    try:
                        cur.execute("SELECT Threshold from sfx where id=" + str(sfxID[0]))
                        threshold = cur.fetchone()[0]
                    except Exception as err:
                        logger.info('Select threshold from sfx error.')
                        logger.error(err)
                        continue
                    if cmpInterval(threshold, str(YVI[0]) + "." + str(YVI[1]) + "." + str(YVI[2])):
                        isSupport = "Supported"
                        try:
                            cur.execute("INSERT INTO support(sfxID, scopusID, batchID) values(" + str(sfxID[0]) + "," + str(scopusID) + "," + str(batchID) + ")")
                            if debug:
                                print ("INSERT INTO support(sfxID, scopusID, batchID) values(" + str(sfxID[0]) + "," + str(scopusID) + "," + str(batchID) + ")")
                            conn.commit()
                        except Exception as err:
                            logger.info('Create relation in support error.')
                            logger.error(err)
                            continue
                        # 蒐集sfx所對應的主題id
                        try:
                            cur.execute("SELECT tid from relation_sfx_theme where sfxid = " + str(sfxID[0]))
                            resultOfTid = cur.fetchall()
                            for Tid in resultOfTid:
                                themeIDStr += str(Tid[0])
                                themeIDStr += ","
                        except Exception as err:
                            logger.info('Search themeID error.')
                            logger.error(err)
                            continue
                        # 判別是否有付費
                        try:
                            cur.execute("SELECT isfree from sfx where id = " + str(sfxID[0]))
                            if cur.fetchone()[0] == 0:
                                isPaid = "Paid"
                            elif isPaid == "":
                                isPaid = "nPaid"
                        except Exception as err:
                            logger.info('paid error.')
                            logger.error(err)
                            continue

                    else:
                        if debug:
                            print (threshold)
                            print (YVI)
                            print(colored('Interval not match.', 'red'))

            else:
                if debug:
                    print(row[8].value)
                    print(colored('ISSN not match.', 'red'))
                outputFile.write(isSupport)
                outputFile.write('\n')
                continue

            outputFile.write(isSupport)
            if themeIDStr is not "":
                if debug:
                    print(colored('Match.', 'yellow'))
                outputFile.write('\t')
                outputFile.write(isPaid)
                outputFile.write('\t')
                outputResult(themeIDStr)
            outputFile.write('\n')
    else:
        if debug:
            print('Please Input the File.')
        return
    print (str(batchID))
    outputFile.close()


def insertDB(row):
    sqlStmt = "INSERT INTO scopus(author_keyword, book_name, year, source_name, volume, issue, DOI, link, ISSN, ISBN) values("
    valStr = ""
    for col in row:
        if col.value is None:
            valStr += 'NULL'
        else:
            valStr += "\""
            valStr += str(col.value).replace("\"", "'").replace("\\", "")
            valStr += "\""
        valStr += ", "
    valStr = valStr[0: len(valStr) - 2]
    valStr += ")"
    try:
        cur.execute(sqlStmt + valStr)
        if debug:
            print (sqlStmt + valStr)
        conn.commit()
        # 修改卷期非數字的問題
        scoupusID = cur.lastrowid
    except Exception as err:
        if debug:
            print(sqlStmt + valStr)
        logger.error(err)
        conn.rollback()
        return -1
    modifyYVI(scoupusID)
    return scoupusID


def modifyYVI(scoupusID):
    try:
        cur.execute("SELECT year, volume, issue from scopus where id = " + str(scoupusID))
        result = cur.fetchone()
    except Exception as err:
        logger.info("Select YVI to modify error.")
        logger.error(err)
        return False
    updateFlag = 0
    tmpList = [result[0], result[1], result[2]]
    if result[0] is None or not re.match('^[0-9]+$', result[0]):
        tmpList[0] = 0
        updateFlag = 1
    if result[1] is None or not re.match('^[0-9]+$', result[1]):
        tmpList[1] = 0
        updateFlag = 1
    if result[2] is None or not re.match('^[0-9]+$', result[2]):
        tmpList[2] = 0
        updateFlag = 1
    if updateFlag:
        try:
            cur.execute("UPDATE scopus set year=" + str(tmpList[0]) + ",volume=" + str(tmpList[1]) + ", issue=" + str(tmpList[2]) + " where id = " + str(scoupusID))
            if debug:
                print ("UPDATE scopus set year=" + str(tmpList[0]) + ",volume=" + str(tmpList[1]) + ", issue=" + str(tmpList[2]) + " where id = " + str(scoupusID))
            conn.commit()
        except Exception as err:
            logger.info('Update YVI error.')
            logger.error(err)
            return False

def outputResult(themeIDList):
    # 尋找主題
    themeStr = ""
    try:
        cur.execute("SELECT name from theme where tid in (" + themeIDList[0: len(themeIDList) - 1] + ")")
        resultOfTheme = cur.fetchall()
        for theme in resultOfTheme:
            themeStr += theme[0]
            themeStr += "|"
        themeStr = themeStr[0: len(themeStr) - 1]
        if debug:
            print(themeStr)
        outputFile.write(themeStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search theme error.')
        logger.error(err)
    # 尋找主題所對應的科系
    departmentStr = ""
    collegeID = ""
    try:
        cur.execute(
            "SELECT name, cid from department where did in (select did from relation_theme_department where tid in (" + themeIDList[
                                                                                                                        0: len(
                                                                                                                            themeIDList) - 1] + "))")
        resultOfDepart = cur.fetchall()
        for depart in resultOfDepart:
            collegeID += str(depart[1])
            collegeID += ","
            departmentStr += depart[0]
            departmentStr += "|"
        departmentStr = departmentStr[0: len(departmentStr) - 1]
        if debug:
            print(departmentStr)
        outputFile.write(departmentStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search department error.')
        logger.error(err)
    # 尋找科系所對應的院
    collegeStr = ""
    try:
        cur.execute("SELECT name from college where cid in (" + collegeID[0: len(collegeID) - 1] + ")")
        resultOfCollege = cur.fetchall()
        for college in resultOfCollege:
            collegeStr += college[0]
            collegeStr += "|"
        collegeStr = collegeStr[0: len(collegeStr) - 1]
        if debug:
            print(collegeStr)
        outputFile.write(collegeStr)
        outputFile.write('\t')
    except Exception as err:
        logger.info('Search college error.')
        logger.error(err)

if __name__ == '__main__':
    # main(filename="../Data/scopus/scopus.xlsx")
    main()