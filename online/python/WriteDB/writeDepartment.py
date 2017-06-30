from openpyxl import load_workbook
import DBconfig as DBconfig
import mysql.connector
import logging
import logging.config


""" 程式用途為更新Department Table """
""" 查資料庫中存不存在，不存在就往後加上去，要對應到院 """

debug = 1
logging.config.fileConfig("../logger.conf")
logger = logging.getLogger("root")
try:
    conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                   host=DBconfig.host)
    cur = conn.cursor()
except Exception as err:
    logger.error(err)

def writeDepartment(filename = ""):
    if filename == "":
        return False

    wb = load_workbook(filename=filename)
    ws = wb[wb.sheetnames[0]]

    for row in ws['A2:D' + str(ws.max_row)]:
        depName = row[3].value
        if not checkIfExist(depName=depName):
            colName = row[1].value
            if not checkIfExist(colName=colName):
                stmt = "Insert into college(name) values(\"" + colName + "\")"
                if debug:
                    print (stmt)
                try:
                    cur.execute(stmt)
                    conn.commit()
                except Exception as err:
                    logger.error(err)
            stmt = "Select cid from college where name = \"" + colName + "\""
            if debug:
                print (stmt)
            try:
                cur.execute(stmt)
                cid = cur.fetchone()[0]
            except Exception as err:
                logger.error(err)

            stmt = "Insert into department(name,cid) values(\"" + depName + "\", " + str(cid) + ")"
            if debug:
                print (stmt)
            try:
                cur.execute(stmt)
                conn.commit()
            except Exception as err:
                logger.error(err)

def checkIfExist(depName="", colName=""):
    if depName is not "":
        stmt = "Select count(*) from department where name = \"" + depName + "\""
    elif colName is not "":
        stmt = "Select count(*) from college where name = \"" + colName + "\""
    if debug:
        print (stmt)
    try:
        cur.execute(stmt)
        return cur.fetchone()[0]
    except Exception as err:
        logger.error(err)
        return
if __name__ == '__main__':
    # writeDepartment(filename="./院系表20170630.xls")
    writeDepartment(filename="./testdata.xlsx")