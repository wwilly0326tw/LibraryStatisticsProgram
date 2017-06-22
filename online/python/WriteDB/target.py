from openpyxl import load_workbook
import Program.DBconfig as DBconfig
import mysql.connector
import sys

"""此程式用於寫入Target欄位"""

try:
    conn = mysql.connector.connect(user=DBconfig.user, password=DBconfig.password, database=DBconfig.database,
                                   host=DBconfig.host)
    cur = conn.cursor()
except Exception as err:
    sys.exit(-1)


wb = load_workbook(filename='../../Data/SFX/target.xlsx')
ws = wb[wb.sheetnames[0]]
dataStr = 'A1:A' + str(ws.max_row)
for row in ws[dataStr]:
    targetName = row[0].value
    print ("Insert into target values('" + targetName + "')")
    cur.execute("Insert into target(name) values(\"" + targetName + "\")")
    conn.commit()